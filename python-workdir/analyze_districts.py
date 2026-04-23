from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin, urlparse

try:
    from urllib3.exceptions import NotOpenSSLWarning

    warnings.filterwarnings("ignore", category=NotOpenSSLWarning)
except Exception:
    pass

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

DISTRICT_HEADER_HINTS = (
    "school district name",
    "district name",
    "district",
    "school district",
    "name",
)
WEBSITE_HEADER_HINTS = (
    "website link",
    "website",
    "web site",
    "url",
    "link",
)

REQUEST_TIMEOUT_SECONDS = 12
MAX_INTERNAL_PAGES = 20
MAX_TEXT_LENGTH = 10000
EXCLUDED_SCHEMES = {"mailto", "tel", "javascript"}
HEADER_FILL = PatternFill("solid", fgColor="123A59")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HYPERLINK_FONT = Font(color="0563C1", underline="single")
CELL_BORDER = Border(
    left=Side(style="thin", color="D9E0E7"),
    right=Side(style="thin", color="D9E0E7"),
    top=Side(style="thin", color="D9E0E7"),
    bottom=Side(style="thin", color="D9E0E7"),
)
BOARDDOCS_PATTERN = re.compile(r"https?://[^\s\"'<>)]*boarddocs[^\s\"'<>)]*", re.IGNORECASE)
BOARDDOCS_TERMS = ("boarddocs",)
BROAD_DOC_TERMS = (
    "agenda",
    "agendas",
    "minutes",
    "meeting",
    "meetings",
    "board of education",
    "school board",
    "governance",
    "board meeting",
    "board meetings",
    "board packet",
    "meeting packet",
    "board agenda",
)
POLICY_TERMS = ("policy", "policies", "regulation", "regulations")
NAVIGATION_TERMS = (
    "board",
    "boarddocs",
    "agenda",
    "agendas",
    "minutes",
    "meeting",
    "meetings",
    "governance",
    "trustees",
    "school board",
    "board of education",
    "policies",
    "documents",
)
GENERAL_NAVIGATION_TERMS = (
    "district information",
    "about",
    "about us",
    "our district",
    "departments",
    "administration",
    "resources",
    "quick links",
    "site map",
    "sitemap",
    "school board",
    "board of school directors",
    "board meeting",
)
BOARDDOCS_HOST_HINTS = ("boarddocs.com",)
BOARDDOCS_STRONG_MEETING_TERMS = (
    "boarddocs® meeting",
    "search meetings",
    "search agenda",
    "include attachments",
)
BOARDDOCS_COMBINED_MEETING_TERMS = (
    ("meetings", "agenda"),
    ("meetings", "minutes"),
    ("agenda", "minutes"),
    ("featured", "meetings"),
)
BOARDDOCS_POLICY_ONLY_TERMS = (
    "boarddocs® policy",
    "policy book",
    "policy status",
    "search policies",
)


@dataclass
class DistrictInput:
    district_name: str
    website_link: str


@dataclass
class DistrictResult:
    district_name: str
    website_link: str
    boarddocs_link: str
    is_boarddocs: bool
    reason: str
    confidence: str


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, dest="input_path")
    parser.add_argument("--output", required=True, dest="output_path")
    args = parser.parse_args()

    input_path = Path(args.input_path)
    output_path = Path(args.output_path)

    districts = load_districts(input_path)
    session = build_session()

    boarddocs_rows: list[DistrictResult] = []
    non_boarddocs_rows: list[DistrictResult] = []
    review_rows: list[DistrictResult] = []

    for item in districts:
        result = analyze_district(session, item)
        if result.confidence == "low":
            review_rows.append(result)
        if result.is_boarddocs:
            boarddocs_rows.append(result)
        else:
            non_boarddocs_rows.append(result)

    export_workbook(output_path, boarddocs_rows, non_boarddocs_rows, review_rows)

    payload = {
        "summary": {
            "totalDistricts": len(districts),
            "boardDocsCount": len(boarddocs_rows),
            "nonBoardDocsCount": len(non_boarddocs_rows),
        }
    }
    print(json.dumps(payload, ensure_ascii=True))
    return 0


def load_districts(input_path: Path) -> list[DistrictInput]:
    rows = read_rows(input_path)
    if not rows:
        raise ValueError("The input workbook is empty.")

    header_index = find_header_row(rows)
    header_row = [normalize_cell(value) for value in rows[header_index]]
    district_col = find_column_index(header_row, DISTRICT_HEADER_HINTS)
    website_col = find_column_index(header_row, WEBSITE_HEADER_HINTS)

    if district_col is None:
        raise ValueError(
            "Unable to find a district name column. Add a header such as 'School District Name' or 'District'."
        )

    districts: list[DistrictInput] = []
    for row in rows[header_index + 1 :]:
        district_name = normalize_cell(get_cell(row, district_col))
        website_link = normalize_cell(get_cell(row, website_col)) if website_col is not None else ""

        if district_name:
            districts.append(
                DistrictInput(
                    district_name=district_name,
                    website_link=website_link,
                )
            )

    if not districts:
        raise ValueError("No usable district rows were found below the header row.")

    return districts


def read_rows(input_path: Path) -> list[list[object]]:
    suffix = input_path.suffix.lower()

    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [row for row in csv.reader(handle)]

    workbook = load_workbook(input_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def find_header_row(rows: list[list[object]]) -> int:
    best_index = 0
    best_score = -1

    for index, row in enumerate(rows[:10]):
        normalized = " ".join(normalize_cell(value).lower() for value in row if normalize_cell(value))
        score = score_header_candidates(normalized)
        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def score_header_candidates(text: str) -> int:
    score = 0
    if any(term in text for term in DISTRICT_HEADER_HINTS):
        score += 2
    if any(term in text for term in WEBSITE_HEADER_HINTS):
        score += 2
    return score


def find_column_index(headers: list[str], hints: Iterable[str]) -> int | None:
    best_index = None
    best_score = -1

    for index, header in enumerate(headers):
        normalized = header.lower()
        score = 0
        for hint in hints:
            if hint == normalized:
                score += 3
            elif hint in normalized:
                score += 1

        if score > best_score:
            best_index = index
            best_score = score

    return best_index if best_score > 0 else None


def get_cell(row: list[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=2,
        backoff_factor=0.4,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "HEAD"),
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            )
        }
    )
    return session


def analyze_district(session: requests.Session, district: DistrictInput) -> DistrictResult:
    website_url = normalize_url(district.website_link)

    if not website_url:
        return DistrictResult(
            district_name=district.district_name,
            website_link=district.website_link,
            boarddocs_link="",
            is_boarddocs=False,
            reason="Missing website link",
            confidence="low",
        )

    home_page = fetch_page(session, website_url)
    if not home_page:
        return DistrictResult(
            district_name=district.district_name,
            website_link=website_url,
            boarddocs_link="",
            is_boarddocs=False,
            reason="Website could not be reached",
            confidence="low",
        )

    evidence_segments = [home_page["title"], home_page["text"]]
    match = discover_boarddocs_link(session, home_page, evidence_segments)

    if not match:
        return DistrictResult(
            district_name=district.district_name,
            website_link=home_page["url"],
            boarddocs_link="",
            is_boarddocs=False,
            reason="No BoardDocs evidence found",
            confidence="low",
        )

    boarddocs_assessment = assess_boarddocs_site(session, match["url"])

    if boarddocs_assessment["page"]:
        evidence_segments.extend(
            [
                boarddocs_assessment["page"]["title"],
                boarddocs_assessment["page"]["text"],
            ]
        )

    combined_evidence = " ".join(segment for segment in evidence_segments if segment).lower()
    site_mentions_board_materials = any(term in combined_evidence for term in BROAD_DOC_TERMS)

    is_boarddocs = boarddocs_assessment["is_meeting_site"] or (
        boarddocs_assessment["confidence"] == "medium" and site_mentions_board_materials
    )
    reason = boarddocs_assessment["reason"]

    return DistrictResult(
        district_name=district.district_name,
        website_link=home_page["url"],
        boarddocs_link=boarddocs_assessment["resolved_url"] or match["url"],
        is_boarddocs=is_boarddocs,
        reason=reason,
        confidence=str(boarddocs_assessment["confidence"]),
    )


def normalize_url(value: str) -> str:
    cleaned = normalize_cell(value)
    if not cleaned:
        return ""

    if not cleaned.startswith(("http://", "https://")):
        cleaned = f"https://{cleaned}"

    return cleaned


def assess_boarddocs_site(session: requests.Session, boarddocs_url: str) -> dict[str, object]:
    best_non_meeting_match: dict[str, object] | None = None

    for candidate_url in build_boarddocs_candidate_urls(boarddocs_url):
        page = fetch_page(session, candidate_url)
        if not page:
            continue

        score = score_boarddocs_page(page)
        if score["is_meeting_site"]:
            return {
                "is_meeting_site": True,
                "confidence": score["confidence"],
                "reason": score["reason"],
                "resolved_url": page["url"],
                "page": page,
            }

        if score["is_policy_only"]:
            best_non_meeting_match = {
                "is_meeting_site": False,
                "confidence": score["confidence"],
                "reason": score["reason"],
                "resolved_url": page["url"],
                "page": page,
            }

    if best_non_meeting_match:
        return best_non_meeting_match

    return {
        "is_meeting_site": False,
        "confidence": "low",
        "reason": "BoardDocs link found but no public meeting navigation was detected",
        "resolved_url": boarddocs_url,
        "page": None,
    }


def build_boarddocs_candidate_urls(boarddocs_url: str) -> list[str]:
    candidates = [boarddocs_url]
    parsed = urlparse(boarddocs_url)
    host = parsed.netloc.lower()

    if not any(hint in host for hint in BOARDDOCS_HOST_HINTS):
        return candidates

    path = parsed.path
    marker = "/Board.nsf/"

    if marker in path:
        prefix = path.split(marker, 1)[0]
        public_root = f"{parsed.scheme}://{parsed.netloc}{prefix}{marker}Public?open"
        candidates.append(public_root)

    deduped: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        deduped.append(candidate)

    return deduped


def score_boarddocs_page(page: dict[str, str]) -> dict[str, object]:
    title = normalize_cell(page["title"]).lower()
    title_and_url = normalize_cell(f"{page['title']} {page['url']}").lower()
    control_labels = extract_control_labels(page["html"])
    control_text = " ".join(sorted(control_labels))
    combined = f"{title_and_url} {control_text}"

    positive_hits = [term for term in BOARDDOCS_STRONG_MEETING_TERMS if term in combined]
    combined_hits = [
        "/".join(term_pair)
        for term_pair in BOARDDOCS_COMBINED_MEETING_TERMS
        if all(term in combined for term in term_pair)
    ]
    policy_hits = [term for term in BOARDDOCS_POLICY_ONLY_TERMS if term in combined]
    has_meetings = "meetings" in control_labels or "meeting" in control_labels
    has_agenda = "agenda" in control_labels or "agendas" in control_labels
    has_minutes = "minutes" in control_labels or "minute" in control_labels
    has_policy = "policies" in control_labels or "policy" in control_labels
    has_search_meetings = "search meetings" in combined
    has_search_policies = "search policies" in combined
    has_lt_title = "boarddocs® lt" in title or "boarddocs lt" in title
    has_pl_title = "boarddocs® pl" in title or "boarddocs pl" in title

    if "boarddocs® meeting" in combined or has_lt_title or has_search_meetings:
        return {
            "is_meeting_site": True,
            "is_policy_only": False,
            "confidence": "high",
            "reason": "BoardDocs page exposes public meetings",
        }

    if has_meetings:
        return {
            "is_meeting_site": True,
            "is_policy_only": False,
            "confidence": "high",
            "reason": "BoardDocs navigation includes a Meetings section",
        }

    if positive_hits and (has_agenda or has_minutes or combined_hits):
        return {
            "is_meeting_site": True,
            "is_policy_only": False,
            "confidence": "high",
            "reason": "BoardDocs page shows meetings, agenda, or minutes navigation",
        }

    if has_agenda and has_minutes and not has_pl_title:
        return {
            "is_meeting_site": True,
            "is_policy_only": False,
            "confidence": "medium",
            "reason": "BoardDocs page shows public meeting-related navigation",
        }

    if has_pl_title:
        return {
            "is_meeting_site": False,
            "is_policy_only": True,
            "confidence": "high",
            "reason": "BoardDocs title indicates a policy library",
        }

    if has_policy and not has_meetings and not has_lt_title and not has_search_meetings:
        return {
            "is_meeting_site": False,
            "is_policy_only": True,
            "confidence": "high",
            "reason": "BoardDocs appears to expose policies only",
        }

    return {
        "is_meeting_site": False,
        "is_policy_only": False,
        "confidence": "low",
            "reason": "BoardDocs page did not clearly expose a Meetings section",
    }


def extract_control_labels(html: str) -> set[str]:
    soup = BeautifulSoup(html, "html.parser")
    labels: set[str] = set()

    for element in soup.find_all(["a", "button", "input", "label", "option"]):
        text = normalize_cell(element.get_text(" ", strip=True))
        if text:
            labels.add(text.lower())

        for attribute in ("aria-label", "title", "value", "placeholder", "alt"):
            attribute_value = normalize_cell(element.get(attribute, ""))
            if attribute_value:
                labels.add(attribute_value.lower())

    return labels


def fetch_page(session: requests.Session, url: str) -> dict[str, str] | None:
    candidates = [url]
    if url.startswith("https://"):
        candidates.append(f"http://{url.removeprefix('https://')}")

    for candidate in candidates:
        try:
            response = session.get(candidate, timeout=REQUEST_TIMEOUT_SECONDS, allow_redirects=True)
            response.raise_for_status()
            content_type = response.headers.get("content-type", "").lower()
            if "text/html" not in content_type and "application/xhtml+xml" not in content_type:
                continue

            soup = BeautifulSoup(response.text, "html.parser")
            title = normalize_cell(soup.title.get_text(" ", strip=True) if soup.title else "")
            page_text = normalize_cell(soup.get_text(" ", strip=True))[:MAX_TEXT_LENGTH]

            return {
                "url": response.url,
                "html": response.text,
                "title": title,
                "text": page_text,
            }
        except requests.RequestException:
            continue

    return None


def discover_boarddocs_link(
    session: requests.Session,
    home_page: dict[str, str],
    evidence_segments: list[str],
) -> dict[str, str] | None:
    queue = [{"url": home_page["url"], "context": home_page["url"], "priority": 100, "page": home_page}]
    queue.extend(build_fallback_seed_candidates(home_page["url"]))
    visited: set[str] = set()

    while queue and len(visited) < MAX_INTERNAL_PAGES:
        queue.sort(key=lambda item: item["priority"], reverse=True)
        current = queue.pop(0)
        current_url = current["url"]

        if current_url in visited:
            continue

        visited.add(current_url)
        current_page = current.get("page") or fetch_page(session, current_url)
        if not current_page:
            continue

        evidence_segments.extend([current.get("context", current_url), current_page["title"]])
        direct_match, candidate_links = inspect_page(current_page["url"], current_page["html"], current.get("context", ""))
        if direct_match:
            return direct_match

        for candidate in candidate_links:
            if candidate["url"] in visited:
                continue
            queue.append(candidate)

    return None


def build_fallback_seed_candidates(home_url: str) -> list[dict[str, object]]:
    parsed = urlparse(home_url)
    if not parsed.scheme or not parsed.netloc:
        return []

    base_root = f"{parsed.scheme}://{parsed.netloc}"
    candidates = [
        "/departments",
        "/en-US/departments",
        "/en/departments",
        "/district-information",
        "/en-US/district-information",
        "/school-board",
        "/en-US/school-board",
        "/board",
        "/en-US/board",
        "/site-map",
        "/sitemap",
    ]

    seeded: list[dict[str, object]] = []
    seen: set[str] = set()
    for path in candidates:
        url = f"{base_root}{path}"
        if url in seen:
            continue
        seen.add(url)
        seeded.append(
            {
                "url": url,
                "context": path.strip("/") or home_url,
                "priority": 55,
            }
        )

    return seeded


def inspect_page(page_url: str, html: str, inherited_context: str = "") -> tuple[dict[str, str] | None, list[dict[str, object]]]:
    soup = BeautifulSoup(html, "html.parser")
    direct_match = extract_direct_match(page_url, html, inherited_context)
    candidates: list[dict[str, object]] = []

    if direct_match:
        return direct_match, candidates

    for anchor in soup.find_all("a", href=True):
        href = normalize_cell(anchor.get("href", ""))
        if not href:
            continue

        parsed = urlparse(href)
        if parsed.scheme and parsed.scheme.lower() in EXCLUDED_SCHEMES:
            continue

        resolved = urljoin(page_url, href)
        anchor_text = normalize_cell(anchor.get_text(" ", strip=True))
        searchable_text = f"{anchor_text} {resolved}".lower()

        if any(term in searchable_text for term in BOARDDOCS_TERMS):
            return (
                {
                    "url": resolved,
                    "context": anchor_text or inherited_context or page_url,
                },
                candidates,
            )

        crawl_priority = get_crawl_priority(page_url, resolved, searchable_text)
        if crawl_priority > 0:
            candidates.append(
                {
                    "url": resolved,
                    "context": anchor_text or inherited_context or resolved,
                    "priority": crawl_priority,
                }
            )

    deduped: list[dict[str, object]] = []
    seen: set[str] = set()
    for candidate in candidates:
        if candidate["url"] in seen:
            continue
        seen.add(candidate["url"])
        deduped.append(candidate)

    return None, deduped


def extract_direct_match(page_url: str, html: str, inherited_context: str) -> dict[str, str] | None:
    match = BOARDDOCS_PATTERN.search(html)
    if not match:
        return None

    raw_url = match.group(0)
    resolved = urljoin(page_url, raw_url)
    return {
        "url": resolved,
        "context": inherited_context or page_url,
    }


def get_crawl_priority(source_url: str, target_url: str, searchable_text: str) -> int:
    source_host = urlparse(source_url).netloc.lower()
    target_host = urlparse(target_url).netloc.lower()
    same_site = not target_host or hosts_are_related(source_host, target_host)

    if not same_site:
        return 0

    priority = 0
    if any(term in searchable_text for term in NAVIGATION_TERMS):
        priority += 100
    if any(term in searchable_text for term in GENERAL_NAVIGATION_TERMS):
        priority += 40

    target_path = urlparse(target_url).path.lower()
    if target_path and target_path not in ("/",):
        priority += 10
    if any(segment in target_path for segment in ("/live-feed", "/article", "/site-map", "/documents")):
        priority += 20

    if "#" in target_url or target_url.rstrip("/") == source_url.rstrip("/"):
        priority -= 100

    return priority


def hosts_are_related(source_host: str, target_host: str) -> bool:
    source_host = source_host.split(":")[0]
    target_host = target_host.split(":")[0]

    if source_host == target_host:
        return True

    source_core = strip_common_subdomain_prefix(source_host)
    target_core = strip_common_subdomain_prefix(target_host)

    return (
        source_core == target_core
        or source_host.endswith(f".{target_core}")
        or target_host.endswith(f".{source_core}")
    )


def strip_common_subdomain_prefix(host: str) -> str:
    prefixes = ("www.", "app.", "cms.")
    for prefix in prefixes:
        if host.startswith(prefix):
            return host[len(prefix) :]
    return host


def export_workbook(
    output_path: Path,
    boarddocs_rows: list[DistrictResult],
    non_boarddocs_rows: list[DistrictResult],
    review_rows: list[DistrictResult],
) -> None:
    workbook = Workbook()
    boarddocs_sheet = workbook.active
    boarddocs_sheet.title = "BoardDocs Districts"
    non_boarddocs_sheet = workbook.create_sheet("Non BoardDocs Districts")
    review_sheet = workbook.create_sheet("Needs Review")

    write_sheet(
        boarddocs_sheet,
        ["School District Name", "Website Link", "BoardDocs Link", "Detection Reason", "Confidence"],
        [
            [row.district_name, row.website_link, row.boarddocs_link, row.reason, row.confidence]
            for row in boarddocs_rows
        ],
    )
    write_sheet(
        non_boarddocs_sheet,
        ["School District Name", "Website Link", "Detection Reason", "Confidence"],
        [
            [row.district_name, row.website_link, row.reason, row.confidence]
            for row in non_boarddocs_rows
        ],
    )
    write_sheet(
        review_sheet,
        ["School District Name", "Website Link", "BoardDocs Link", "Detection Reason", "Confidence"],
        [
            [row.district_name, row.website_link, row.boarddocs_link, row.reason, row.confidence]
            for row in review_rows
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_sheet(sheet, headers: list[str], rows: list[list[str]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER

            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.font = HYPERLINK_FONT

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        adjusted_width = min(max(max_length + 2, 18), 55)
        sheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # noqa: BLE001
        print(str(error), file=sys.stderr)
        raise
